import json
import shutil
from tqdm import tqdm
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import datetime
from csv import DictReader
from shutil import copyfile, move
import time
import pyautogui
import os


def create_output_folder(path):
    if not os.path.isdir(path):
        os.mkdir(path)


def remove_output_folder(path):
    if os.path.isdir(path):
        shutil.rmtree(path)
    else:
        pass

remove_output_folder("input/positions")
print('done')
create_output_folder("input/positions")
with open("positions.json", "r") as f:
    data = json.load(f)

for k, v in tqdm(data.items()):

    template_wb = load_workbook("NEW_TEMPLATE.xlsx")
    #

    details_sheet = template_wb['details']

    position = k
    industry = v["parent_industry"]
    sub_industry = v["sub_industry"]

    details_sheet["A2"] = position
    details_sheet["B2"] = industry
    details_sheet["C2"] = sub_industry

    #

    intros = v["intro"]

    intros_sheet = template_wb['intros']
    counter = 1
    for a in intros:
        counter += 1
        ref = "A" + str(counter)
        intros_sheet[ref] = a['text']

    #

    bodies = v["body"]
    bodies_sheet = template_wb['bodies']

    counter = 1
    for b in bodies:
        counter += 1

        id_ref = "A" + str(counter)
        def_resp_ref = "B" + str(counter)
        custom_ref = "C" + str(counter)
        custom_one_ref = "D" + str(counter)
        question_ref = "E" + str(counter)
        fillers_ref = "F" + str(counter)
        type_ref = "G" + str(counter)

        bodies_sheet[id_ref] = b['id']
        bodies_sheet[def_resp_ref] = b['default_response']
        bodies_sheet[custom_ref] = b['custom_text_response']
        bodies_sheet[custom_one_ref] = b['custom_text_response_one_element']
        bodies_sheet[question_ref] = b['question']
        if b['fillers']:
            fillers_str = ','.join(b['fillers'])
        else:
            fillers_str = ""
        bodies_sheet[fillers_ref] = fillers_str
        bodies_sheet[type_ref] = b['type']

    #

    outros = v["outro"]
    outros_sheet = template_wb['outros']

    counter = 1
    for c in outros:
        counter += 1
        ref = "A" + str(counter)
        outros_sheet[ref] = c['text']

    #

    file_name = "hand written position files/" + position + ".xlsx"

    template_wb.save(file_name)
