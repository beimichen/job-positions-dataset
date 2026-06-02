# import Workbook
import json

from openpyxl import load_workbook

# # create Workbook object
# wb = Workbook()
# # set file path
# filepath = "/home/ubuntu/demo.xlsx"
# # save workbook
# wb.save(filepath)

import os

from tqdm import tqdm

from upload_to_s3 import upload_positions_json


def get_details(dest_path):
    wb = load_workbook(dest_path)
    sheet = wb['details']
    position_name = sheet['A2'].value
    parent_industry = sheet['B2'].value
    sub_industry = sheet['C2'].value

    return position_name, parent_industry, sub_industry


def get_intros(dest_path):
    wb = load_workbook(dest_path)
    sheet = wb['intros']
    intros = []
    for i in sheet['A']:
        if i.value:
            if i.value == "intros":
                pass
            else:
                intros.append(i.value)
    return intros


def convert_intros_to_obj(intros):
    obj = []
    for i in intros:
        dictionary = {'text': i}
        obj.append(dictionary)
    return obj


def get_bodies(dest_path):
    workbook = load_workbook(dest_path)
    sheet = workbook['bodies']
    bodies_list = []
    counter = 0
    for i in sheet['A']:
        counter += 1
        obj = {}
        if i.value:
            if i.value == 'id':
                pass
            else:
                id_ref = 'A' + str(counter)
                id = sheet[id_ref].value
                if id:
                    obj['id'] = int(id)
                else:
                    break

                default_response_ref = 'B' + str(counter)
                default_response = sheet[default_response_ref].value
                if default_response:
                    obj['default_response'] = default_response
                else:
                    obj['default_response'] = ""

                custom_text_response_ref = 'C' + str(counter)
                custom_text_response = sheet[custom_text_response_ref].value
                if custom_text_response:
                    obj['custom_text_response'] = custom_text_response
                else:
                    obj['custom_text_response'] = ""

                custom_text_response_one_element_ref = 'D' + str(counter)
                custom_text_response_one_element = sheet[custom_text_response_one_element_ref].value
                if custom_text_response_one_element:
                    obj['custom_text_response_one_element'] = custom_text_response_one_element
                else:
                    obj['custom_text_response_one_element'] = ""

                question_ref = 'E' + str(counter)
                question = sheet[question_ref].value
                if question:
                    obj['question'] = question
                else:
                    obj['question'] = ""

                fillers_ref = 'F' + str(counter)
                fillers = sheet[fillers_ref].value
                if fillers:
                    obj['fillers'] = fillers
                else:
                    obj['fillers'] = None

                type_ref = 'G' + str(counter)
                type_out = sheet[type_ref].value
                if type_out:
                    obj['type'] = type_out
                else:
                    obj['type'] = ""

                bodies_list.append(obj)

    return bodies_list


def get_outros(dest_path):
    wb = load_workbook(dest_path)
    sheet = wb['outros']
    outros_to_compile = []
    for i in sheet['A']:
        if i.value:
            if i.value == "outros":
                pass
            else:
                outros_to_compile.append(i.value)
    return outros_to_compile


def convert_outros_to_obj(outros_to_compile):
    obj = []
    for i in outros_to_compile:
        dictionary = {'text': i}
        obj.append(dictionary)
    return obj

def process_excel_to_json():
    data = {}

    for filename in tqdm(os.listdir(r"input/templates")):
        if filename.endswith(".xlsx"):
            dest_path = os.path.join(r"input/templates", filename)
            # prep data
            position, industry, sub_industry = get_details(dest_path)
            print(position)
            print(filename)
            clean_filename = filename.replace(".xlsx","")
            if clean_filename == position:
                pass
            else:
                print("fix position in xslx: ",filename)
            intros = get_intros(dest_path)
            intros_obj = convert_intros_to_obj(intros)
            data[position] = {}
            data[position]['intro'] = intros_obj

            bodies = get_bodies(dest_path)
            data[position]['body'] = bodies

            outros = get_outros(dest_path)
            outros_obj = convert_outros_to_obj(outros)
            data[position]['outro'] = outros_obj

            data[position]['parent_industry'] = industry
            data[position]['sub_industry'] = sub_industry
        else:
            print("passing")
            pass

    for filename in tqdm(os.listdir(r"input/positions")):
        if filename.endswith(".xlsx"):
            dest_path = os.path.join(r"input/positions", filename)
            # prep data
            position, industry, sub_industry = get_details(dest_path)
            print(position)
            print(filename)
            clean_filename = filename.replace(".xlsx","")
            if clean_filename == position:
                pass
            else:
                print("fix position in xslx: ",filename)
            intros = get_intros(dest_path)
            intros_obj = convert_intros_to_obj(intros)
            data[position] = {}
            data[position]['intro'] = intros_obj

            bodies = get_bodies(dest_path)
            data[position]['body'] = bodies

            outros = get_outros(dest_path)
            outros_obj = convert_outros_to_obj(outros)
            data[position]['outro'] = outros_obj

            data[position]['parent_industry'] = industry
            data[position]['sub_industry'] = sub_industry
        else:
            print("passing")
            pass


    with open(r'data/positions.json', 'w') as outfile:
        json.dump(data, outfile, sort_keys=True, indent=2)

    # read file
    with open(r'data/positions.json', 'r') as myfile:
        data=myfile.read()
        obj = json.loads(data)
        print(obj)
        for k,v in obj.items():
            print(k)
            print(v["body"])
            all_body = v["body"]
            for h in all_body:
                type_body = h["type"]
                if type_body:
                    pass
                else:
                    input("FAIL")

        print("passed")

